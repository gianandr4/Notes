import win32com.client
import sys
import os
import time
import statistics as st
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

def calc(x):
    x=time.strftime('%H:%M:%S', time.gmtime(x))
    return(x)

###############################################################################
outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
inbox = outlook.GetDefaultFolder(6).Folders.Item("DIA")
messages = inbox.Items#-----------------------------------------------INIT
messages.Sort("[ReceivedTime]", True)
message = messages.GetLast()#-----------------------------------------INIT
# TOTAL NUMBER OF EMAILS IN THE FOLDER
print("Total number",len(inbox.items)) 
input("Press Enter to start...")
start_time = time.time()  
#MENUMENUMENUMENUMENUMENUMENUMENUMENUMENUMENUMENUMENUMENUMENUMENUMENUMENUMENUMENU
print("DIA \n----\n")
input("From date(dd/mm/yyyy): ")
messages = messages.Restrict("[ReceivedTime] >= '" +frodate.format(datetime,'%m/%d/%Y %H:%M %p')+"'")
input("To date(dd/mm/yyyy): ")
messages = messages.Restrict("[ReceivedTime] <= '" +todate.format(datetime,'%m/%d/%Y %H:%M %p')+"'")
print("Total messages: ",len(messages),'from: ',frodate,'to',todate,'\n\n')

#############################################################################

fax=['N','P','E']
data=[]#counts SRs
b=False
synd=0
apos=0
tot=0
counter=['','','','','']
dups=[]
SRs=''
faxs=''
faxss=''
SRss=''
faxss=''
faxsss=''
SRsss=''
faxssss=''
SRssss=''
###############################################################################
# CORE CODE
messages.Sort("[ReceivedTime]", False)
for i,message in enumerate(messages):
    try:        
        if (str(message.sender) != "Collection_Business_Strategy, GRC02" and
            str(message.sender) !="Nikitopoulos, Athanasios" and
            str(message.sender) !="Barba, Apostolia"):continue
            
        if (message.subject.startswith('Προσωρινές διακοπές DIA') or 
            message.subject.startswith('Προσωρινή διακοπή DIA') or 
            message.subject.startswith('RE: Προσωρινές διακοπές DIA')):
            for line in str.split(message.body):
                if len(line) == 10 and (line[2] in fax or line.isdigit()):#FAX
                    faxs=line
                if len(line) >= 13 and line[:2]=='1-':#SR
                    SRs=line
                if len(faxs)==10 and len(SRs) >=13:
                    dups.append([faxs,SRs,message.receivedtime,'0','0','0'])
                    faxs=''
                    SRs=''  
                    

        elif (message.subject.startswith('Eπανασύνδεση DIA') or
            message.subject.startswith(' ΕΠΑΝΑΣΥΝΔΕΣΗ  DIA') or
            message.subject.startswith('Eπανασυνδέσεις DIA') or 
            message.subject.startswith('RE: Eπανασύνδεση DIA') or 
            message.subject.startswith('Eπανασύνδεση Dia')):
            for line in str.split(message.body):
                if len(line) == 10 and (line[2] in fax or line.isdigit()):#FAX
                    faxss=line
                if len(line) >= 13 and line[:2]=='1-':#SR
                    SRss=line
                if len(faxss)==10 and len(SRss) >=13:
                    for i,dd in enumerate (dups):
                        t=False
                        if faxss in dd:
                            t=True
                            if dups[i][3] != '' and dups[i][4] != '':
                                temp=dups[i]
                                dups[i][3]=SRss #check if empty first
                                dups[i][4]=message.receivedtime
                    if t==False:
                        dups.append([faxss,'0','0',SRss,message.receivedtime,'0'])
                        #else: CREATE A LIST OF FAULTS
                    SRss=''
                    faxss=''
                    
        elif message.subject.startswith('ΠΡΟΣΩΡΙΝΗ ΔΙΑΚΟΠΗ ETHERNET'):
            for line in str.split(message.body):
                if len(line) == 32 and line[10]=='-' and line[21]=='/':head,sep,faxsss=line.partition('/')#FAX
                if len(line) >= 13 and line[:2]=='1-':SRsss=line#SR
                if len(faxsss)==10 and len(SRsss) >=13:
                    dups.append([faxsss,SRsss,message.receivedtime,'0','0','0'])
                    faxsss=''
                    SRsss=''
                    
        elif message.subject.startswith('Επανασύνδεση ETHERNET'):
#            print(message.body)
            for line in str.split(message.body):
                if len(line) == 32 and line[10]=='-' and line[21]=='/':head,sep,faxssss=line.partition('/')#FAX
                if len(line) >= 13 and line[:2]=='1-':SRssss=line#SR
                if len(faxssss)==10 and len(SRssss) >=13:
                    for i,dd in enumerate (dups):
                        if faxssss in dd:
                            t=False
                            if dups[i][3] != '0' and dups[i][4] != '0':
                                t=True
                                dups[i][3]=SRssss #check if empty first
                                dups[i][4]=message.receivedtime
                    if not t:
                        dups.append([faxssss,'0','0',SRssss,message.receivedtime,'0'])
                    SRss=''
                    faxss=''
                          
        else:
            data.append([message.receivedtime,message.sender,message.subject])
    except:
        print('Failed to retrieve',message.receivedtime,'-',message.subject)
###############################################################################   
#df0=pd.DataFrame(dups)
ddi=dups.copy()
# DATA MANAGEMENT 
times=[]#CONTAINS THE DIFFERENCE OF HOURS           
for i,d in enumerate(dups):

    if dups[i][4]=='0' or dups[i][2]=='0':
        dups[i][2]=str(dups[i][2]).split('+',1)[0]
        dups[i][4]=str(dups[i][4]).split('+',1)[0]

#        dups[i][3]='Disconnected'
#        dups[i][4]='Disconnected'
        dups[i][5]='Disconnected'
        
    else:

        dups[i][5]=dups[i][4] - dups[i][2]
        times.append(float(dups[i][5].total_seconds()))#TIME DIFFERENCE IN INT, SECONDS
        dups[i][2]=str(dups[i][2]).split('+',1)[0]
        dups[i][4]=str(dups[i][4]).split('+',1)[0]
        dups[i][5]=str(dups[i][5]).split('+',1)[0]

###############################################################################   
## STATISTICS
print('STATISTICS')
try:
    if len(times)>0: 
        print('Mean Value:',calc(st.mean(times)))
        print('Median Value:',calc(st.median(times)))
        print('Harmonic Mean Value:',calc(st.harmonic_mean(times)))
        print('Median Grouped Value:',calc(st.median_grouped(times)))
        print('Standard Deviation:',calc(st.stdev(times)))
        print('Population Standard Devviation:',calc(st.pstdev(times)))
except:print('no')


df0=pd.DataFrame(dups)
df=pd.DataFrame(dups)
print('here')

df.columns=['FAX','Disconnect_SR','Disconnect_TIME','Reconnect_SR','Reconnect_TIME','DC_DURATION']
df=df.sort_values(by=['DC_DURATION'])
df=df.reset_index(drop=True)
df1=df[df['Reconnect_SR']!='Disconnected']
df2=df[df['Reconnect_SR']=='Disconnected']
df3=pd.DataFrame(times)
print('TABLE\n',df.to_string())


if len(data)>0:
    for i in range(len(data)):
        for j in range(len(data[i])):
            data[i][j]=str(data[i][j])
    print("\n----ERRORS----:")
    data=pd.DataFrame(data)
    print(data.to_string())      
    
del_char=['[',']',':','*','?','/']

name=str(frodate+'-'+todate)
name=name.replace('/','')
name=name.replace(':','')

df.sort_values(by=['DC_DURATION'])
file_name=input('Type file name: ')
df.to_excel(excel_writer=file_name+'.xlsx',engine='xlsxwriter',sheet_name=name)

#
#
#workbook = xlsxwriter.Workbook(file_name+'.xlsx')
#worksheet = workbook.add_worksheet()
#worksheet.write('A1', 'Item', bold)
#worksheet.write('B1', 'Date', bold)
#worksheet.write('C1', 'Cost', bold)
#




print('Total: ',len(df[df['Disconnect_SR']!='0'])+len(df[df['Reconnect_SR']!='0']))
print ('Disconnected: ',len(df[df['Disconnect_SR']!='0']))
print ('Reconnected: ',len(df[df['Reconnect_SR']!='0']))
print('Still Disconnected: ', (len(df[df['DC_DURATION']=='Disconnected'])))






df1=pd.DataFrame({'Total':[len(df[df['Disconnect_SR']!='0'])+len(df[df['Reconnect_SR']!='0'])],
                  'Disconnected':[len(df[df['Disconnect_SR']!='0'])],
                  'Reconnected':[len(df[df['Reconnect_SR']!='0'])],
                  'Still Disconnected':[len(df[df['DC_DURATION']=='Disconnected'])]})

book = load_workbook(file_name+'.xlsx')
writer = pd.ExcelWriter(file_name+'.xlsx', engine = 'openpyxl')
writer.book = book
#df.to_excel(writer, sheet_name = 'Sum') 


df1.to_excel(writer, sheet_name = 'Sum')
writer.save()
writer.close()

#
#
#for message in messages:
#    print(message.subject)
#    if (message.subject.startswith('Προσωρινές διακοπές DIA') or 
#        message.subject.startswith('Προσωρινή διακοπή DIA') or 
#        message.subject.startswith('RE: Προσωρινές διακοπές DIA')): 
#        print('nai')
#
#
#
#
#
#
#
#
#
#
#




